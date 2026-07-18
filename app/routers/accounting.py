from fastapi import APIRouter, Depends, HTTPException, Response, File, UploadFile, Request
from sqlmodel import Session, select
from app.core.database import get_session
from app.core.security import get_current_admin, get_current_user
from app.services.accounting_service import create_massive_debt, process_payment, migrate_historical_debt
from app.services.notification_service import send_push_to_section
from app.services.report_service import generate_individual_debt_pdf
from app.core.limiter import limiter
from app.models.entities import Debt, Resident
from pydantic import BaseModel
from typing import List

router = APIRouter(prefix="/api/v1/accounting", tags=["Accounting"])

class MassiveDebtRequest(BaseModel):
    concepto: str
    monto: float
    seccion_id: int = None

class PaymentRequest(BaseModel):
    residente_id: int
    deuda_id: int
    monto_entregado: float

@router.post("/massive-debt")
def generate_massive_debt(
    data: MassiveDebtRequest,
    session: Session = Depends(get_session),
    admin: Resident = Depends(get_current_admin)
):
    count = create_massive_debt(session, data.concepto, data.monto, data.seccion_id)

    # Notificar a los residentes sobre la nueva deuda
    titulo = f"Nueva Cuota: {data.concepto}"
    mensaje = f"Se ha generado un cobro de ${data.monto} por concepto de {data.concepto}. Por favor revisa tu sección de deudas."
    send_push_to_section(session, data.seccion_id, titulo, mensaje)

    return {"message": f"Deuda generada y notificada a {count} residentes."}

@router.post("/pay")
@limiter.limit("10/minute")
def pay_debt(request: Request, data: PaymentRequest, session: Session = Depends(get_session)):
    payment = process_payment(session, data.residente_id, data.deuda_id, data.monto_entregado)
    if not payment:
        raise HTTPException(status_code=400, detail="Error procesando el pago o deuda ya pagada.")
    return payment

@router.get("/my-debts", response_model=List[Debt])
def get_my_debts(current_user: Resident = Depends(get_current_user), session: Session = Depends(get_session)):
    statement = select(Debt).where(Debt.residente_id == current_user.id, Debt.pagada == False)
    return session.exec(statement).all()

@router.get("/resident/{resident_id}/debts", response_model=List[Debt])
def get_resident_debts(resident_id: int, admin: Resident = Depends(get_current_admin), session: Session = Depends(get_session)):
    statement = select(Debt).where(Debt.residente_id == resident_id)
    return session.exec(statement).all()

@router.get("/resident/{resident_id}/pdf-report")
def download_resident_pdf(resident_id: int, session: Session = Depends(get_session)):
    file_data = generate_individual_debt_pdf(session, resident_id)
    if not file_data:
        raise HTTPException(status_code=404, detail="Residente no encontrado")
    return Response(
        content=file_data,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=deuda_residente_{resident_id}.pdf"}
    )

@router.post("/migrate-excel")
async def migrate_excel(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    admin: Resident = Depends(get_current_admin)
):
    """
    Endpoint para subir un Excel y cargar deudas históricas.
    Formato esperado: Columnas [Casa, Monto, Concepto]
    """
    import io
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="La librería 'openpyxl' no está instalada en el servidor.")

    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    sheet = wb.active

    results = []
    # Saltamos la primera fila (encabezados)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        casa, monto, concepto = row[0], row[1], row[2]

        if casa and monto:
            res = migrate_historical_debt(
                session=session,
                identificador_unidad=str(casa),
                concepto=str(concepto or "Saldo Inicial Migración"),
                monto_total=float(monto)
            )
            results.append({"casa": casa, "status": res})

    return {"processed": len(results), "details": results}
